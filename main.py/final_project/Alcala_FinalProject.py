import tkinter as tk
import openpyxl as op
from tkinter import ttk,messagebox



def medecines():

    def display():
        workbook = op.load_workbook("MedicinesDB.xlsx")
        sheet = workbook.active


        for content in table.get_children():
            table.delete(content)

        for rows in sheet.iter_rows(min_row=2,values_only=True):
            table.insert("",tk.END,values=rows)

        workbook.save("MedicinesDB.xlsx")
    
    def validation():
        name = name_entry.get()
        gen_name = gen_name_entry.get()
        dosage = dosage_entry.get()
        category = category_entry.get()
        expiration = exp_entry.get()

        if not name or not gen_name or not dosage or not category or not expiration:
            messagebox.showerror("Error","All fields are required", parent=med_win)
            return False
        return True

    def create():
        if not validation():
            return 

        
        name = name_entry.get()
        gen_name = gen_name_entry.get()
        dosage = dosage_entry.get()
        category = category_entry.get()
        expiration = exp_entry.get()

        workbook = op.load_workbook("MedicinesDB.xlsx")
        sheet = workbook.active

        new_id = sheet.max_row

        sheet.append([new_id,name,gen_name,dosage,category,expiration])
        workbook.save("MedicinesDB.xlsx")

        messagebox.showinfo("Succesful","Record Added Successfully", parent=med_win)
        display()
        
        # Clear entry fields after successful save
        name_entry.delete(0, tk.END)
        gen_name_entry.delete(0, tk.END)
        dosage_entry.delete(0, tk.END)
        category_entry.delete(0, tk.END)
        exp_entry.delete(0, tk.END)

    def auto_populate(event):
        selected = table.focus()
        values = table.item(selected,"values")

        if values:
            name_entry.delete(0,tk.END)
            gen_name_entry.delete(0,tk.END)
            category_entry.delete(0,tk.END)
            dosage_entry.delete(0,tk.END)
            exp_entry.delete(0,tk.END)

            name_entry.insert(0,values[1])
            gen_name_entry.insert(0,values[2])
            category_entry.insert(0,values[4])
            dosage_entry.insert(0,values[3])
            exp_entry.insert(0,values[5])

    def update():
        selected = table.focus()
        if not selected:
            messagebox.showinfo("Error","Select a record first", parent=med_win)
            return
        
        if not validation():
            return
        
        name = name_entry.get()
        gen_name = gen_name_entry.get()
        category = category_entry.get()
        dosage = dosage_entry.get()
        expiration = exp_entry.get()

        values = table.item(selected,"values")
        record_id = values[0]

        workbook = op.load_workbook("MedicinesDB.xlsx")
        sheet = workbook.active

        for rows in sheet.iter_rows(min_row = 2):
            if str(record_id) == str(rows[0].value):
                rows[1].value = name
                rows[2].value = gen_name
                rows[4].value = category
                rows[3].value = dosage
                rows[5].value = expiration
                workbook.save("MedicinesDB.xlsx")
                messagebox.showinfo("Succes","Record updated successfully", parent=med_win)
                break
        display()
         # Clear entry fields after successful save
        name_entry.delete(0, tk.END)
        gen_name_entry.delete(0, tk.END)
        dosage_entry.delete(0, tk.END)
        category_entry.delete(0, tk.END)
        exp_entry.delete(0, tk.END)

    def delete():

        selected = table.focus()

        if not selected:
            messagebox.showerror("Error","Select First",parent=med_win)
            
        confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete?",parent=med_win)
        if not confirm:
            return

        values = table.item(selected,"values")
        record_id = values[0]

        workbook = op.load_workbook("MedicinesDB.xlsx")
        sheet = workbook.active

        for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
            if str(record_id) == str(row[0].value):
                sheet.delete_rows(i)
                break

        # if deleted iorder
        new_id = 1
        for row in sheet.iter_rows(min_row=2):
            row[0].value = new_id
            new_id += 1

        workbook.save("MedicinesDB.xlsx")
        messagebox.showinfo("Success","Successfully Deleted",parent=med_win)
        display()
        
        # Clear entry fields after successful delete
        name_entry.delete(0, tk.END)
        gen_name_entry.delete(0, tk.END)
        dosage_entry.delete(0, tk.END)
        category_entry.delete(0, tk.END)
        exp_entry.delete(0, tk.END)


           
           
       

            



    med_win = tk.Toplevel(window)
    med_win.title("Medecines")
    med_win.geometry("1015x530")
    med_win.configure(bg="#5F8D7E")
    for i in range(4):
        med_win.grid_columnconfigure(i, weight=1)

    # # Make this popup modal and keep dialogs attached to it
    # med_win.transient(window)
    # med_win.grab_set()
    # med_win.focus_set()

    med_ti = tk.Label(med_win,text="Medecines Information Form",font=("Times New Roman",30,"bold"),fg="white",bg="#5F8D7E")
    med_ti.grid(row=1,column=0,columnspan=4,pady=(10,5),sticky="ew")

    name_entry = tk.Entry(med_win, font=("Poppins", 12))
    name_entry.grid(row=3, column=0, padx=10, pady=(10, 0), sticky="ew")

    gen_name_entry = tk.Entry(med_win, font=("Poppins", 12))
    gen_name_entry.grid(row=3, column=1, padx=10, pady=(10, 0), sticky="ew")

    dosage_entry = tk.Entry(med_win, font=("Poppins", 12))
    dosage_entry.grid(row=3, column=2, padx=10, pady=(10, 0), sticky="ew")

    category_entry = tk.Entry(med_win, font=("Poppins", 12))
    category_entry.grid(row=3, column=3, padx=10, pady=(10, 0), sticky="ew")

    name_label = tk.Label(med_win, text="Name", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    name_label.grid(row=4, column=0, padx=10, pady=(2, 10), sticky="w")

    gen_name_label = tk.Label(med_win, text="Generic Name", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    gen_name_label.grid(row=4, column=1, padx=10, pady=(2, 10), sticky="w")

    dosage_label = tk.Label(med_win, text="Dosage", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    dosage_label.grid(row=4, column=2, padx=10, pady=(2, 10), sticky="w")

    category_label = tk.Label(med_win, text="Category", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    category_label.grid(row=4, column=3, padx=10, pady=(2, 10), sticky="w")

    exp_entry = tk.Entry(med_win, font=("Poppins", 12))
    exp_entry.grid(row=5, column=0, columnspan=2, padx=10, pady=(10, 0), sticky="ew")

    exp_label = tk.Label(med_win, text="Expiration Date", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    exp_label.grid(row=6, column=0, columnspan=2, padx=10, pady=(0, 10), sticky="w")

    submit_btn = tk.Button(med_win, text="Submit", font=("Poppins", 12, "bold"), bg="#2A7F62",command=create)
    submit_btn.grid(row=5, column=2, padx=10, pady=30, sticky="ew")

    update_btn = tk.Button(med_win, text="Update",font=("Poppins", 12, "bold"), bg="#4A90E2",command=update)
    update_btn.grid(row=5, column=3, padx=10, pady=30, sticky="ew")

    delete_btn = tk.Button(med_win, text="Delete", bg="#D9534F", fg="white",font=("Poppins", 12, "bold"),command=delete)
    delete_btn.grid(row=6, column=2, columnspan=2, padx=10, pady=(0,30), sticky="ew")

    # Table
    table = ttk.Treeview(med_win,columns=("ID", "Name", "Generic Name", "Dosage", "Category", "Expiration Date"),show="headings")

    for headings in ("ID", "Name", "Generic Name", "Dosage", "Category", "Expiration Date"):
        table.heading(headings, text=headings)

    table.column("ID", width=80)
    table.column("Name", width=200)
    table.column("Generic Name", width=200)
    table.column("Dosage", width=170)
    table.column("Category", width=170)
    table.column("Expiration Date", width=170)
    table.grid(row=2, column=0, columnspan=5, padx=10, pady=10)
    table.bind("<<TreeviewSelect>>",auto_populate)
    display()

def inventory():

    def display():

        workbook = op.load_workbook("InventoryDB.xlsx")
        sheet = workbook.active

        for content in table.get_children():
            table.delete(content)

        for rows in sheet.iter_rows(min_row=2,values_only=True):
            table.insert("",tk.END,values=rows)

            workbook.save("InventoryDB.xlsx")

    def validation():
        medicines = medicines_entry.get()
        qty = qty_entry.get()
        date = date_entry.get()


        if not medicines or not qty or not date:
            messagebox.showerror("Error","All fields are required",parent=inv_win)
            return False

        if not qty.isdigit():
            messagebox.showerror("Error","Quantity should be a number",parent=inv_win)
            return False
        return True

    def create():
        if not validation():
            return

        medicines = medicines_entry.get()
        qty = int(qty_entry.get())
        date = date_entry.get()

       
        
        if qty <= 20:
            status = "▼ Low Stock"
        else:
            status = "▲ High Stock"

        workbook = op.load_workbook("InventoryDB.xlsx")
        sheet = workbook.active

        
        new_id = sheet.max_row

        sheet.append([new_id,medicines,qty,status,date])
        workbook.save("InventoryDB.xlsx")

        messagebox.showinfo("Successfull","Record added successfully",parent=inv_win)
        display()

        medicines_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        date_entry.delete(0,tk.END)

    def auto_populate(event):

        selected = table.focus()
        values = table.item(selected,"values")

        if selected:
            medicines_entry.delete(0,tk.END)
            qty_entry.delete(0,tk.END)
            date_entry.delete(0,tk.END)

            medicines_entry.insert(0,values[1])
            qty_entry.insert(0,values[2])
            date_entry.insert(0,values[4])

    def update():
        selected = table.focus()

        if not selected:
            messagebox.showerror("Error","Select a record first",parent=inv_win)
            return

        if not validation():
            return
        
        medicines = medicines_entry.get()
        qty = qty_entry.get()
        date = date_entry.get()

        

        values = table.item(selected,"values")
        record_id = values[0]

        workbook = op.load_workbook("InventoryDB.xlsx")
        sheet = workbook.active

        for rows in sheet.iter_rows(min_row=2):
            if str(record_id) == str(rows[0].value):
                rows[1].value = medicines
                rows[2].value = qty
                rows[4].value = date
                workbook.save("InventoryDB.xlsx")
                messagebox.showinfo("Success","Record updated successfully",parent=inv_win)
                break
        display()

        medicines_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        date_entry.delete(0,tk.END)
    
    def delete():
        selected = table.focus()
        
        if not selected:
            messagebox.showerror("Error","Select first",parent=inv_win)
            return

        confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete?",parent=inv_win)
        if not confirm:
            return
        
        values = table.item(selected,"values")

        record_id = values[0]

        workbook = op.load_workbook("InventoryDB.xlsx")
        sheet = workbook.active

        for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
            if str(record_id) == str(row[0].value):
                sheet.delete_rows(i)
                break

        new_id = 1
        for row in sheet.iter_rows(min_row=2):
            row[0].value = new_id
            new_id += 1

        workbook.save("InventoryDB.xlsx")
        messagebox.showinfo("Success","Record successfully deleted",parent=inv_win)
        display()

        medicines_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        date_entry.delete(0,tk.END)
            


    inv_win = tk.Toplevel(window)      
    inv_win.title("Medecines")
    inv_win.geometry("1015x530")
    inv_win.configure(bg="#5F8D7E")
    for i in range(3):
        inv_win.grid_columnconfigure(i, weight=1)

    inv_ti = tk.Label(inv_win,text="Medicine Inventory Form",font=("Times New Roman",30,"bold"),fg="white",bg="#5F8D7E")
    inv_ti.grid(row=1, column=0, columnspan=3, pady=(20,10), sticky="ew")

    medicines_entry = tk.Entry(inv_win, font=("Poppins", 12))
    medicines_entry.grid(row=3, column=0, padx=10, pady=(10, 0), sticky="ew")

    date_entry = tk.Entry(inv_win, font=("Poppins", 12))
    date_entry.grid(row=3, column=1, padx=10, pady=(10, 0), sticky="ew")

    qty_entry = tk.Entry(inv_win, font=("Poppins", 12))
    qty_entry.grid(row=3, column=2, padx=10, pady=(10, 0), sticky="ew")

    medicines_label = tk.Label(inv_win, text="Medicines", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    medicines_label.grid(row=4, column=0, padx=10, pady=(2, 10), sticky="w")

    date_label = tk.Label(inv_win, text="Date", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    date_label.grid(row=4, column=1, padx=10, pady=(2, 10), sticky="w")

    qty_label = tk.Label(inv_win, text="Quantity", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    qty_label.grid(row=4, column=2, padx=10, pady=(2, 10), sticky="w")

    submit_btn = tk.Button(inv_win, text="Submit", font=("Poppins", 12, "bold"), bg="#2A7F62",command=create)
    submit_btn.grid(row=5, column=0, padx=10, pady=20, sticky="ew")

    update_btn = tk.Button(inv_win, text="Update",font=("Poppins", 12, "bold"), bg="#4A90E2",command=update)
    update_btn.grid(row=5, column=1, padx=10, pady=20, sticky="ew")

    delete_btn = tk.Button(inv_win, text="Delete", bg="#D9534F", fg="white",font=("Poppins", 12, "bold"),command=delete)
    delete_btn.grid(row=5, column=2, padx=10, pady=20, sticky="ew")

    table = ttk.Treeview(inv_win,columns=("ID","Medicine","Quantity","Status","Expiration Date"),show="headings")

    for headings in ("ID","Medicine","Quantity","Status","Expiration Date"):
        table.heading(headings,text=headings)

    table.grid(row=2, column=0, columnspan=5, padx=10, pady=10)
    table.bind("<<TreeviewSelect>>",auto_populate)
    display()

def sales():

    def display():
        workbook = op.load_workbook("SalesDB.xlsx")
        sheet = workbook.active

        for content in table.get_children():
            table.delete(content)

        for rows in sheet.iter_rows(min_row=2,values_only=True):
            table.insert("",tk.END,values=rows)

        workbook.save("SalesDB.xlsx")

    def validation():

        medicine = medicines_entry.get()
        qty = qty_entry.get()
        price = price_entry.get()

        if not medicine or not qty or not price:
            messagebox.showerror("Error","All fields required",parent=sales_win)
            return False

        if not price.isdigit() or not qty.isdigit():
            messagebox.showerror("Error","Price and Quantity should be a number",parent=sales_win)
            return False
        return True

    def create():
        if not validation():
            return

        medicine = medicines_entry.get()
        qty = int(qty_entry.get())
        price = int(price_entry.get())

        total = qty * price

        workbook = op.load_workbook("SalesDB.xlsx")
        sheet = workbook.active

        new_id = sheet.max_row

        sheet.append([new_id,medicine,qty,price,total])
        workbook.save("SalesDB.xlsx")
        messagebox.showinfo("Success","Record added successfully",parent=sales_win)
        display()

        medicines_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        price_entry.delete(0,tk.END)

    def auto_populate(event):
        selected = table.focus()
        values = table.item(selected,"values")

        if selected:
            medicines_entry.delete(0,tk.END)
            qty_entry.delete(0,tk.END)
            price_entry.delete(0,tk.END)

            medicines_entry.insert(0,values[1])
            qty_entry.insert(0,values[2])
            price_entry.insert(0,values[3])

    def update():
        selected = table.focus()
        
        if not selected:
            messagebox.showerror("Error","Select first",parent=sales_win)
            return
        if not validation():
            return
        
        medicines = medicines_entry.get()
        qty = int(qty_entry.get())
        price = int(price_entry.get())

        total = qty * price

        values = table.item(selected,"values")
        new_id = values[0]

        workbook = op.load_workbook("SalesDB.xlsx")
        sheet = workbook.active

        for rows in sheet.iter_rows(min_row=2):
            if  str(new_id) == str(rows[0].value):
                rows[1].value = medicines
                rows[2].value = qty
                rows[3].value = price
                rows[4].value = total
                workbook.save("SalesDB.xlsx")
                messagebox.showinfo("Success","Record updated successfully",parent=sales_win)
                break
        display()

        medicines_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        price_entry.delete(0,tk.END)

    def delete():
        selected = table.focus()

        if not selected:
            messagebox.showerror("Error","Select first",parent=sales_win)
            return
        confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete?",parent=sales_win)
        if not confirm:
            return
        values = table.item(selected,"values")
        record_id = values[0]

        workbook = op.load_workbook("SalesDB.xlsx")
        sheet = workbook.active

        for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
            if str(record_id) == str(row[0].value):
                sheet.delete_rows(i)
                break
        
        new_id = 1
        for row in sheet.iter_rows(min_row=2):
            row[0].value = new_id
            new_id += 1

        workbook.save("SalesDB.xlsx")
        messagebox.showinfo("Error","Record deleted successfully",parent=sales_win)
        
        display()
        medicines_entry.delete(0,tk.END)
        qty_entry.delete(0,tk.END)
        price_entry.delete(0,tk.END)




    sales_win = tk.Toplevel(window)
    sales_win.title("Sales")
    sales_win.geometry("1015x530")
    sales_win.configure(bg="#5F8D7E")
    for i in range(3):
        sales_win.grid_columnconfigure(i, weight=1)

    sales_ti = tk.Label(sales_win,text="Sales Entry Form",font=("Times New Roman",30,"bold"),fg="white",bg="#5F8D7E")
    sales_ti.grid(row=1, column=0, columnspan=3, pady=(20,10), sticky="ew")

    medicines_entry = tk.Entry(sales_win, font=("Poppins", 12))
    medicines_entry.grid(row=3, column=0, padx=10, pady=(10, 0), sticky="ew")

    price_entry = tk.Entry(sales_win, font=("Poppins", 12))
    price_entry.grid(row=3, column=1, padx=10, pady=(10, 0), sticky="ew")

    qty_entry = tk.Entry(sales_win, font=("Poppins", 12))
    qty_entry.grid(row=3, column=2, padx=10, pady=(10, 0), sticky="ew")

    medicines_label = tk.Label(sales_win, text="Medicines", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    medicines_label.grid(row=4, column=0, padx=10, pady=(2, 10), sticky="w")

    price_label = tk.Label(sales_win, text="Price", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    price_label.grid(row=4, column=1, padx=10, pady=(2, 10), sticky="w")

    qty_label = tk.Label(sales_win, text="Quantity", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    qty_label.grid(row=4, column=2, padx=10, pady=(2, 10), sticky="w")

    submit_btn = tk.Button(sales_win, text="Submit", font=("Poppins", 12, "bold"), bg="#5CB85C",command=create)
    submit_btn.grid(row=5, column=0, padx=10, pady=20, sticky="ew")

    update_btn = tk.Button(sales_win, text="Update",font=("Poppins", 12, "bold"), bg="#4A90E2",command=update)
    update_btn.grid(row=5, column=1, padx=10, pady=20, sticky="ew")

    delete_btn = tk.Button(sales_win, text="Delete", bg="#D9534F", fg="white",font=("Poppins", 12, "bold"),command=delete)
    delete_btn.grid(row=5, column=2, padx=10, pady=20, sticky="ew")

    table = ttk.Treeview(sales_win,columns=("ID","Medicine","Quantity","Price","Total"),show="headings")

    for headings in ("ID","Medicine","Quantity","Price","Total"):
        table.heading(headings,text=headings)

    table.grid(row=2, column=0, columnspan=5, padx=10, pady=10)
    table.bind("<<TreeviewSelect>>",auto_populate)
    display()

def supplier():

    def display():
        workbook = op.load_workbook("SupplierDB.xlsx")
        sheet = workbook.active

        for content in table.get_children():
            table.delete(content)

        for rows in sheet.iter_rows(min_row=2,values_only=True):
            table.insert("",tk.END,values=rows)

        workbook.save("SupplierDB.xlsx")
    
    def validation():
        name = name_entry.get()
        supplier = supplier_name_entry.get()
        con_no = con_no_entry.get()
        email = email_entry.get()
        address = address_entry.get()

        if not name or not supplier or not con_no or not email or not address:
            messagebox.showerror("Error","All fields are required", parent=supp_win)
            return False
        return True

    def create():
        if not validation():
            return 

        name = name_entry.get()
        supplier = supplier_name_entry.get()
        con_no = con_no_entry.get()
        email = email_entry.get()
        address = address_entry.get()

        workbook = op.load_workbook("SupplierDB.xlsx")
        sheet = workbook.active

        new_id = sheet.max_row

        sheet.append([new_id,name,supplier,con_no,email,address])
        workbook.save("SupplierDB.xlsx")

        messagebox.showinfo("Succesful","Record Added Successfully", parent=supp_win)
        display()
        
        # Clear entry fields after successful save
        name_entry.delete(0, tk.END)
        supplier_name_entry.delete(0, tk.END)
        con_no_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)
        address_entry.delete(0, tk.END)

    def auto_populate(event):
        selected = table.focus()
        values = table.item(selected,"values")

        if values:
            name_entry.delete(0, tk.END)
            supplier_name_entry.delete(0, tk.END)
            con_no_entry.delete(0, tk.END)
            email_entry.delete(0, tk.END)
            address_entry.delete(0, tk.END)

            name_entry.insert(0,values[1])
            supplier_name_entry.insert(0,values[2])
            con_no_entry.insert(0,values[3])
            email_entry.insert(0,values[4])
            address_entry.insert(0,values[5])

    def update():
        selected = table.focus()
        if not selected:
            messagebox.showinfo("Error","Select a record first", parent=supp_win)
            return
        
        if not validation():
            return
        
        name = name_entry.get()
        supplier = supplier_name_entry.get()
        con_no = con_no_entry.get()
        email = email_entry.get()
        address = address_entry.get()

        values = table.item(selected,"values")
        record_id = values[0]

        workbook = op.load_workbook("SupplierDB.xlsx")
        sheet = workbook.active

        for rows in sheet.iter_rows(min_row = 2):
            if str(record_id) == str(rows[0].value):
                rows[1].value = name
                rows[2].value = supplier
                rows[3].value = con_no
                rows[4].value = email
                rows[5].value = address
                workbook.save("SupplierDB.xlsx")
                messagebox.showinfo("Succes","Record updated successfully", parent=supp_win)
                break
        display()

        name_entry.delete(0, tk.END)
        supplier_name_entry.delete(0, tk.END)
        con_no_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)
        address_entry.delete(0, tk.END)

    def delete():

        selected = table.focus()

        if not selected:
            messagebox.showerror("Error","Select First",parent=supp_win)
            return

        confirm = messagebox.askyesnocancel("Confirm","Are you sure you want to delete?",parent=supp_win)
        if not confirm:
            return

        values = table.item(selected,"values")
        record_id = values[0]

        workbook = op.load_workbook("SupplierDB.xlsx")
        sheet = workbook.active

        for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
            if str(record_id) == str(row[0].value):
                sheet.delete_rows(i)
                break

        # if deleted iorder
        new_id = 1
        for row in sheet.iter_rows(min_row=2):
            row[0].value = new_id
            new_id += 1

        workbook.save("SupplierDB.xlsx")
        messagebox.showinfo("Success","Successfully Deleted",parent=supp_win)
        display()
        
        # Clear entry fields after successful delete
        name_entry.delete(0, tk.END)
        supplier_name_entry.delete(0, tk.END)
        con_no_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)
        address_entry.delete(0, tk.END)

    supp_win = tk.Toplevel(window)
    supp_win.title("Medecines")
    supp_win.geometry("1015x530")
    supp_win.configure(bg="#5F8D7E")
    for i in range(4):
        supp_win.grid_columnconfigure(i, weight=1)

    
    supp_ti = tk.Label(supp_win,text="Supplier Information Form",font=("Times New Roman",30,"bold"),fg="white",bg="#5F8D7E")
    supp_ti.grid(row=1,column=0,columnspan=4,pady=(10,5), sticky="ew")

    name_entry = tk.Entry(supp_win, font=("Poppins", 12))
    name_entry.grid(row=3, column=0, padx=10, pady=(10, 0), sticky="ew")

    supplier_name_entry = tk.Entry(supp_win, font=("Poppins", 12))
    supplier_name_entry.grid(row=3, column=1, padx=10, pady=(10, 0), sticky="ew")

    con_no_entry = tk.Entry(supp_win, font=("Poppins", 12))
    con_no_entry.grid(row=3, column=2, padx=10, pady=(10, 0), sticky="ew")

    email_entry = tk.Entry(supp_win, font=("Poppins", 12))
    email_entry.grid(row=3, column=3, padx=10, pady=(10, 0), sticky="ew")

    name_label = tk.Label(supp_win, text="Company", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    name_label.grid(row=4, column=0, padx=10, pady=(2, 10), sticky="w")

    supp_name_label = tk.Label(supp_win, text="Supplier", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    supp_name_label.grid(row=4, column=1, padx=10, pady=(2, 10), sticky="w")

    con_no_label = tk.Label(supp_win, text="Contact Number", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    con_no_label.grid(row=4, column=2, padx=10, pady=(2, 10), sticky="w")

    email_label = tk.Label(supp_win, text="Email", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    email_label.grid(row=4, column=3, padx=10, pady=(2, 10), sticky="w")

    address_entry = tk.Entry(supp_win, font=("Poppins", 12))
    address_entry.grid(row=5, column=0, columnspan=4, padx=10, pady=(10, 0), sticky="ew")

    address_label = tk.Label(supp_win, text="Address", font=("Poppins", 10, "italic"), bg="#5F8D7E")
    address_label.grid(row=6, column=0, columnspan=4, padx=10, pady=(2, 10), sticky="w")

    submit_btn = tk.Button(supp_win, text="Submit", font=("Poppins", 12, "bold"), bg="#2A7F62",command=create)
    submit_btn.grid(row=7, column=0, padx=10, pady=20, sticky="ew")

    update_btn = tk.Button(supp_win, text="Update",font=("Poppins", 12, "bold"), bg="#4A90E2",command=update)
    update_btn.grid(row=7, column=1, padx=10, pady=20, sticky="ew")

    delete_btn = tk.Button(supp_win, text="Delete", bg="#D9534F", fg="white",font=("Poppins", 12, "bold"),command=delete)
    delete_btn.grid(row=7, column=2, padx=10, pady=20, sticky="ew")

    # Table
    table = ttk.Treeview(supp_win,columns=("ID", "Company", "Supplier Name", "Contact No", "Email", "Address"),show="headings")

    # Table
    table = ttk.Treeview(supp_win,columns=("ID", "Company", "Supplier Name", "Contact No", "Email", "Address"),show="headings")

    for headings in ("ID", "Company", "Supplier Name", "Contact No", "Email", "Address"):
        table.heading(headings, text=headings)

    table.column("ID", width=80)
    table.column("Company", width=180)
    table.column("Supplier Name", width=180)
    table.column("Contact No", width=160)
    table.column("Email", width=180)
    table.column("Address", width=180)
    table.grid(row=2, column=0, columnspan=6, padx=10, pady=10)
    table.bind("<<TreeviewSelect>>",auto_populate)
    display()
    
window = tk.Tk()
window.title("Medecine Management System")
window.geometry('480x520')
window.configure(bg="#5F8D7E")
window.resizable(False, False)
for i in range(6):
    window.grid_columnconfigure(i, weight=1)
    window.grid_rowconfigure(i, weight=0)
window.grid_rowconfigure(0, weight=0)
window.grid_rowconfigure(1, weight=1)

title = tk.Label(window, text="Medecine Management System", font=("Times New Roman",25,"bold"), fg="white", bg="#5F8D7E")
title.grid(row=0, column=0, columnspan=6, pady=20, padx=20, sticky="ew")

frame = tk.Frame(window, bg="#F6F4EB", width=340, height=420)
frame.grid(row=1, column=0, columnspan=6, padx=20, pady=10, sticky="n")
frame.grid_propagate(False)
frame.grid_columnconfigure(0, weight=1)

mede_btn = tk.Button(frame, text="Medecines", font=("Aerial", 16, "bold"), fg="white", bg="#6B8CAE", width=20, height=2, activebackground="#56718D", bd=0, highlightthickness=0, relief="flat",command=medecines)
mede_btn.grid(row=0, column=0, padx=10, pady=(10, 15), sticky="ew")

inven_btn = tk.Button(frame, text="Inventory", font=("Aerial", 16, "bold"), fg="white", bg="#6B8CAE", width=20, height=2, activebackground="#56718D", bd=0, highlightthickness=0, relief="flat",command=inventory)
inven_btn.grid(row=1, column=0, padx=10, pady=15, sticky="ew")

sales_btn = tk.Button(frame, text="Sales", font=("Aerial", 16, "bold"), fg="white", bg="#6B8CAE", width=20, height=2, activebackground="#56718D", bd=0, highlightthickness=0, relief="flat",command=sales)
sales_btn.grid(row=2, column=0, padx=10, pady=15, sticky="ew")

supp_btn = tk.Button(frame, text="Suppliers", font=("Aerial", 16, "bold"), fg="white", bg="#6B8CAE", width=20, height=2, activebackground="#56718D", bd=0, highlightthickness=0, relief="flat",command=supplier)
supp_btn.grid(row=3, column=0, padx=10, pady=(15,10), sticky="ew")

window.mainloop()