import tkinter as tk
import openpyxl as op
from tkinter import ttk,messagebox

def display():
    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active
    
    #clear table 
    for content in tree.get_children():
        tree.delete(content)

    for rows in sheet.iter_rows(min_row=2,values_only=True):
        tree.insert("",tk.END,values=rows)

def validation():
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()

    if not first or not middle or not last or not birth:
        messagebox.showerror("Error","All fieds are required")
        return False
    if not birth.isdigit():
        messagebox.showerror("Error","Birth year should be a number")
        return False
    return True
def create():
    if not validation ():
        return
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = int(birth_entry.get())

    age = 2026 - birth
    workbook = op.load_workbook('excelDB.xlsx')
    sheet = workbook.active

    new_id = sheet.max_row

    sheet.append([new_id,last,first,middle,birth,age])
    workbook.save("excelDB.xlsx")

    messagebox.showinfo("Succes","Record added successfully")
    display()

def auto_populate(event):
    selected = tree.focus()
    values = tree.item(selected,"values")

    if values:
        lname_entry.delete(0,tk.END)
        fname_entry.delete(0,tk.END)
        mname_entry.delete(0,tk.END)
        birth_entry.delete(0,tk.END)

        lname_entry.insert(0,values[1])
        fname_entry.insert(0,values[2])
        mname_entry.insert(0,values[3])
        birth_entry.insert(0,values[4])

def update():

    selected = tree.focus()
    if not selected():
        messagebox.showinfo("Error","Select a record first")
    
    if not validation():
        return
    
    first = fname_entry.get()
    middle = mname_entry.get()
    last = lname_entry.get()
    birth = birth_entry.get()

    age = 2026 - birth

    values = tree.item(selected,"values")
    record_id = values[0]

    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active
    
    for rows in sheet.iter_rows(main_row = 2):
        if str(record_id) == str(rows[0].value):
            rows[1].value = last
            rows[2].value = first
            rows[3].value = middle
            rows[4].value = birth
            rows[5].value = age
        workbook.save("excelDB.xlsx")
        messagebox.showinfo("Succes","Record updated successfully")
        display()

def delete(): 
    selected = tree.focus()

    if not selected:
        messagebox.showerror("Error","Select a record first")
    
    confirm = messagebox.askyesnocancel("Confirm","Are you sure do you want to delete this?")
    if not confirm:
        return
    
    values = tree.item(selected,"values")
    record_id = values[0]

    workbook = op.load_workbook("excelDB.xlsx")
    sheet = workbook.active

    for i,row in enumerate(sheet.iter_rows(min_row=2),start=2):
        if str(record_id) == str(row[0].value):
            sheet.delete_rows(i)
            break
    workbook.save("excelDB.xlsx")
    messagebox.showinfo("Success","Record deleted")

window=tk.Tk()
window.title("Age Calculator")
window.configure(bg="lightgreen")


#Form Title
title = tk.Label ( window, text="Profile Builder", font=("Times New Roman",14,"bold"),bg="lightgreen")
title.grid(row=0, column=0, columnspan=6)

#Frame
genframe = tk.Frame(window,bg="lightgreen",bd=2, relief="groove")
genframe.grid(row=1,column=0, columnspan=6,padx=10,pady=10)

#First Name Entry
fname_entry = tk.Entry(genframe, font=("Poppins",12))
fname_entry.grid(row=2, column=1,columnspan=2,padx=(10,0),pady=(10,0))

fname_label = tk.Label(genframe, text="First Name", font=("Poppins",10,"italic"),bg="lightgreen")
fname_label.grid(row=3, column=1,columnspan=2)

#Middle Name Entry
mname_entry = tk.Entry(genframe, font=("Poppins",12))
mname_entry.grid(row=2, column=3,columnspan=2,padx=(10,0),pady=(10,0))

mname_label = tk.Label(genframe, text="Middle Name", font=("Poppins",10,"italic"),bg="lightgreen")
mname_label.grid(row=3, column=3,columnspan=2)

#Last Name Entry
lname_entry = tk.Entry(genframe, font=("Poppins",12))
lname_entry.grid(row=2, column=5,columnspan=2,padx=(10,10),pady=(10,0))

lname_label = tk.Label(genframe, text="Last Name", font=("Poppins",10,"italic"),bg="lightgreen")
lname_label.grid(row=3, column=5,columnspan=2)

#Birthyear Entry
birth_entry = tk.Entry(genframe, font=("Poppins",12))
birth_entry.grid(row=4, column=1,columnspan=2,padx=(10,0))

birthyear_label = tk.Label(genframe, text="Birth Year", font=("Poppins",10,"italic"),bg="lightgreen")
birthyear_label.grid(row=5, column=2,columnspan=2)

update_btn = tk.Button(window, text="Update",command=update)
update_btn.grid(row=6, column=2)

button= tk.Button(window,text="Submit",font=("Poppins",12,"bold"),bg="lightpink",command=create)
button.grid(row=6, column=0, columnspan=6,pady=(10,20))

delete_btn = tk.Button(window, text="Delete",  bg="red", fg="white",command=delete)
delete_btn.grid(row=6, column=3)

tree = ttk.Treeview(window, columns=("ID","Last","First","Middle","BirthYear","Age"), show="headings")
for col in ("ID","Last","First","Middle","BirthYear","Age"):
    tree.heading(col, text=col)
tree.grid(row=7, column=0, columnspan=4)
tree.bind("<<TreeviewSelect>>",auto_populate)

display()
window.mainloop()